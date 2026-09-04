"""
data/descripcion_campos.py
==========================
Genera la hoja "Diccionario de Datos" que se agrega al Excel del
**Informe completo** del Autodiagnóstico de Calidad (CNBPEO), documentando
las hojas "Autodiagnóstico", "Plan de acción", "Estadísticas" y "Puntaje por
nivel" de ese mismo archivo (ver views/autodiagnostico.py::_informe_completo).

Por qué existe este módulo
---------------------------
La ONE, a través de "Lineamientos y Recomendaciones para Documentar el
Diccionario de Datos" (Dirección de Normativas y Metodologías), clasifica como
"Diccionario de Datos Pasivo" cualquier conjunto de datos que no vive dentro
de un sistema gestor de base de datos, como un archivo XLSX — exactamente el
caso del Informe completo — y exige documentarlo de forma manual. Este módulo
es esa documentación, generada en código para que nunca quede desactualizada
respecto a las columnas realmente exportadas.

NOTA: la versión anterior de este archivo documentaba el módulo viejo de
Indicadores / Fuentes / Factibilidad (SIDOE), que ya no existe. Se reescribió
para el sistema actual (autodiagnóstico CNBPEO: GEI · GPE · GRE).
"""

from __future__ import annotations

import datetime as _dt

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Catálogo de metadatos por columna, agrupado por hoja del Informe completo.
# Las CLAVES son los nombres de columna tal como aparecen en el Excel
# (nombres ya "bonitos", porque los DataFrames se exportan con esas cabeceras).
# Cada entrada: etiqueta, tipo, condicion, dominio, descripcion.
# ---------------------------------------------------------------------------

_META_AUTODIAGNOSTICO: dict[str, dict[str, str]] = {
    "Código": dict(
        etiqueta="Código del elemento", tipo="Texto", condicion="Obligatorio",
        dominio="Código único (ej. GEI-1.1.1)",
        descripcion="Identificador único del elemento de verificación dentro "
                     "de la matriz del autodiagnóstico.",
    ),
    "Nivel": dict(
        etiqueta="Nivel de gestión", tipo="Texto", condicion="Obligatorio",
        dominio="GEI | GPE | GRE",
        descripcion="Nivel de gestión al que pertenece el elemento (GEI = "
                     "Entorno Institucional, GPE = Proceso Estadístico, "
                     "GRE = Resultados Estadísticos).",
    ),
    "Principio": dict(
        etiqueta="Principio", tipo="Texto", condicion="Obligatorio",
        dominio="N.° y nombre del principio (ej. '1. Independencia profesional')",
        descripcion="Principio de calidad del CNBPEO que agrupa al elemento.",
    ),
    "Requisito": dict(
        etiqueta="Requisito de cumplimiento", tipo="Texto", condicion="Obligatorio",
        dominio="Código del requisito (ej. GEI-1.1)",
        descripcion="Requisito de cumplimiento bajo el cual se clasifica el "
                     "elemento de verificación.",
    ),
    "Elemento": dict(
        etiqueta="Elemento de verificación", tipo="Texto", condicion="Obligatorio",
        dominio="Texto libre",
        descripcion="Texto del elemento concreto que la institución debe "
                     "evaluar y evidenciar.",
    ),
    "¿Cumple?": dict(
        etiqueta="¿Cumple con el elemento?", tipo="Texto", condicion="Opcional",
        dominio="SI | NO | N/A",
        descripcion="Indica si la institución cumple con el elemento "
                     "(N/A = no aplica a la institución).",
    ),
    "Nivel de cumplimiento": dict(
        etiqueta="Nivel de cumplimiento", tipo="Texto", condicion="Opcional",
        dominio="No aplica | No Iniciado | Iniciado | Cumplimiento Parcial | "
                 "Cumplimiento Total",
        descripcion="Grado de avance del cumplimiento del elemento. Solo "
                     "'Cumplimiento Total' suma al puntaje (ver hoja "
                     "Estadísticas).",
    ),
    "Evidencia actual": dict(
        etiqueta="Evidencia actual", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Descripción de la evidencia más reciente que respalda el "
                     "cumplimiento del elemento.",
    ),
    "Enlace actual": dict(
        etiqueta="Enlace / Link de evidencia actual", tipo="Texto (URL)", condicion="Opcional",
        dominio="URL válida (SharePoint, Drive, Web)",
        descripcion="Enlace a la evidencia actual almacenada en línea.",
    ),
    "Archivo actual": dict(
        etiqueta="Archivo de evidencia actual", tipo="Texto", condicion="Opcional",
        dominio="Nombre de archivo adjunto (PDF, imagen, Word, Excel)",
        descripcion="Nombre del archivo de evidencia actual adjuntado al "
                     "elemento.",
    ),
    "Evidencia anterior": dict(
        etiqueta="Evidencia anterior", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Descripción de la evidencia de una revisión anterior, "
                     "para comparar el avance en el tiempo.",
    ),
    "Enlace anterior": dict(
        etiqueta="Enlace / Link de evidencia anterior", tipo="Texto (URL)", condicion="Opcional",
        dominio="URL válida (SharePoint, Drive, Web)",
        descripcion="Enlace a la evidencia anterior almacenada en línea.",
    ),
    "Archivo anterior": dict(
        etiqueta="Archivo de evidencia anterior", tipo="Texto", condicion="Opcional",
        dominio="Nombre de archivo adjunto (PDF, imagen, Word, Excel)",
        descripcion="Nombre del archivo de evidencia anterior adjuntado al "
                     "elemento.",
    ),
    "Comentario": dict(
        etiqueta="Comentario", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Observaciones o aclaraciones adicionales sobre el "
                     "elemento.",
    ),
    "Acción de mejora": dict(
        etiqueta="Acción de mejora", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Acción planificada para cumplir o mejorar el elemento.",
    ),
    "Responsable": dict(
        etiqueta="Responsable", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Persona o área responsable del elemento.",
    ),
    "Fecha": dict(
        etiqueta="Fecha de cumplimiento", tipo="Fecha", condicion="Opcional",
        dominio="AAAA-MM-DD",
        descripcion="Fecha en que se cumplió (o se espera cumplir) el "
                     "elemento.",
    ),
}

_META_PLAN: dict[str, dict[str, str]] = {
    "Requerimiento": dict(
        etiqueta="Requerimiento", tipo="Texto", condicion="Obligatorio",
        dominio="Código + descripción del requerimiento",
        descripcion="Requerimiento del plan de acción; agrupa sus actividades, "
                     "insumos y responsables.",
    ),
    "Acción de mejora": dict(
        etiqueta="Acción de mejora", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Acción concreta que se realizará para atender el "
                     "requerimiento.",
    ),
    "Actividades": dict(
        etiqueta="Actividades", tipo="Texto", condicion="Opcional",
        dominio="Texto libre (una actividad por línea)",
        descripcion="Actividades necesarias para cumplir el requerimiento.",
    ),
    "Insumo": dict(
        etiqueta="Insumo", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Recursos o insumos necesarios para ejecutar las "
                     "actividades.",
    ),
    "Presupuesto": dict(
        etiqueta="Presupuesto", tipo="Texto", condicion="Opcional",
        dominio="Monto o 'N/A'",
        descripcion="Presupuesto estimado para el requerimiento.",
    ),
    "Fecha de cumplimiento": dict(
        etiqueta="Fecha de cumplimiento", tipo="Texto / Fecha", condicion="Opcional",
        dominio="Fecha o plazo (ej. DD/MM/AAAA)",
        descripcion="Fecha o plazo comprometido para cumplir el "
                     "requerimiento.",
    ),
    "Responsable": dict(
        etiqueta="Responsable", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Persona(s) o área(s) responsable(s) de la ejecución.",
    ),
    "Indicador verificable": dict(
        etiqueta="Indicador verificable", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Medio de verificación que permite comprobar el "
                     "cumplimiento del requerimiento.",
    ),
    "Riesgo": dict(
        etiqueta="Riesgo", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Riesgos identificados que podrían afectar el "
                     "cumplimiento.",
    ),
    "Acciones de mitigación": dict(
        etiqueta="Acciones de mitigación", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Acciones previstas para mitigar los riesgos "
                     "identificados.",
    ),
    "Observaciones": dict(
        etiqueta="Observaciones", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Notas adicionales sobre el requerimiento del plan.",
    ),
}

_META_ESTADISTICAS: dict[str, dict[str, str]] = {
    "Nivel": dict(
        etiqueta="Nivel de gestión", tipo="Texto", condicion="Obligatorio",
        dominio="GEI | GPE | GRE",
        descripcion="Nivel de gestión del principio.",
    ),
    "N°": dict(
        etiqueta="Número del principio", tipo="Numérico (entero)", condicion="Obligatorio",
        dominio="Entero ≥ 1",
        descripcion="Número del principio dentro de su nivel.",
    ),
    "Principio": dict(
        etiqueta="Principio", tipo="Texto", condicion="Obligatorio",
        dominio="Texto libre",
        descripcion="Nombre del principio de calidad.",
    ),
    "No aplica": dict(
        etiqueta="Elementos 'No aplica'", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Cantidad de elementos del principio marcados como 'No "
                     "aplica' (no cuentan para el puntaje).",
    ),
    "No Iniciado": dict(
        etiqueta="Elementos 'No Iniciado'", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Cantidad de elementos en estado 'No Iniciado'.",
    ),
    "Iniciado": dict(
        etiqueta="Elementos 'Iniciado'", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Cantidad de elementos en estado 'Iniciado' (1%–49%).",
    ),
    "Cumpl. Parcial": dict(
        etiqueta="Elementos en Cumplimiento Parcial", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Cantidad de elementos en 'Cumplimiento Parcial' "
                     "(50%–99%).",
    ),
    "Cumpl. Total": dict(
        etiqueta="Elementos en Cumplimiento Total", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Cantidad de elementos en 'Cumplimiento Total' (100%). "
                     "Es el único estado que suma al puntaje.",
    ),
    "Elementos": dict(
        etiqueta="Total de elementos", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Total de elementos de verificación del principio.",
    ),
    "% Cumplimiento": dict(
        etiqueta="% de Cumplimiento", tipo="Numérico (decimal)", condicion="Calculado",
        dominio="0 a 100",
        descripcion="Porcentaje de cumplimiento = elementos en 'Cumplimiento "
                     "Total' ÷ elementos que aplican × 100 (igual que la hoja "
                     "RESULTADOS de la matriz oficial).",
    ),
}

_META_PUNTAJE_NIVEL: dict[str, dict[str, str]] = {
    "Nivel": dict(
        etiqueta="Nivel de gestión", tipo="Texto", condicion="Obligatorio",
        dominio="GEI | GPE | GRE | GENERAL",
        descripcion="Nivel de gestión, o 'GENERAL' para el puntaje global de "
                     "la institución.",
    ),
    "Gestión": dict(
        etiqueta="Nombre del nivel", tipo="Texto", condicion="Obligatorio",
        dominio="Texto libre",
        descripcion="Nombre completo del nivel de gestión.",
    ),
    "% Cumplimiento": dict(
        etiqueta="% de Cumplimiento", tipo="Numérico (decimal)", condicion="Calculado",
        dominio="0 a 100",
        descripcion="Puntaje del nivel: promedio del % en 'Cumplimiento "
                     "Total' de sus principios. En la fila GENERAL, el "
                     "puntaje global de la institución.",
    ),
}

_METADATOS_POR_HOJA: dict[str, dict[str, dict[str, str]]] = {
    "Autodiagnóstico": _META_AUTODIAGNOSTICO,
    "Plan de acción": _META_PLAN,
    "Estadísticas": _META_ESTADISTICAS,
    "Puntaje por nivel": _META_PUNTAJE_NIVEL,
}

_COLUMNAS_TABLA = [
    "Hoja", "Nombre de la Variable", "Etiqueta", "Tipo de Dato",
    "Condición", "Dominio / Valores Permitidos", "Descripción",
]


def _fila_para_columna(hoja: str, columna: str) -> dict[str, str]:
    """Devuelve la fila del diccionario para una columna dada, con un fallback
    genérico para columnas que no estén documentadas estáticamente (p. ej. si
    en el futuro se agrega una columna nueva a alguna hoja)."""
    meta = _METADATOS_POR_HOJA.get(hoja, {}).get(columna)
    if meta is None:
        return {
            "Hoja": hoja,
            "Nombre de la Variable": columna,
            "Etiqueta": str(columna).replace("_", " ").strip().title(),
            "Tipo de Dato": "Texto",
            "Condición": "Opcional",
            "Dominio / Valores Permitidos": "—",
            "Descripción": "Columna sin descripción documentada.",
        }
    return {
        "Hoja": hoja,
        "Nombre de la Variable": columna,
        "Etiqueta": meta["etiqueta"],
        "Tipo de Dato": meta["tipo"],
        "Condición": meta["condicion"],
        "Dominio / Valores Permitidos": meta["dominio"],
        "Descripción": meta["descripcion"],
    }


def construir_tabla_descripcion_campos(
    hojas: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    """Arma la tabla del Diccionario de Datos a partir de las columnas
    REALMENTE presentes en cada hoja exportada, en el mismo orden en que
    aparecen.

    ``hojas`` es un dict {nombre_de_hoja: DataFrame}, por ejemplo:
        {"Autodiagnóstico": df_auto, "Plan de acción": df_plan,
         "Estadísticas": df_stats, "Puntaje por nivel": df_niv}
    """
    filas = []
    for hoja, df in hojas.items():
        if df is None:
            continue
        for columna in df.columns:
            # Se omite la columna técnica de "aviso" (hojas exportadas vacías).
            if str(columna).strip().lower() == "aviso":
                continue
            filas.append(_fila_para_columna(hoja, str(columna)))
    return pd.DataFrame(filas, columns=_COLUMNAS_TABLA)


# ---------------------------------------------------------------------------
# Escritura en el Excel: bloque de identificación institucional (formato
# "Ficha técnica" de los Lineamientos ONE) + tabla, con formato openpyxl.
# ---------------------------------------------------------------------------

_FILL_ENCABEZADO = PatternFill("solid", fgColor="0F3A7A")
_FONT_ENCABEZADO = Font(color="FFFFFF", bold=True, name="Arial")
_FONT_TITULO = Font(bold=True, size=13, name="Arial", color="0F3A7A")
_FONT_ETIQUETA_FICHA = Font(bold=True, name="Arial")
_FONT_NORMAL = Font(name="Arial")
_ANCHO_COLUMNAS = [16, 30, 34, 18, 14, 46, 70]


def aplicar_formato_encabezado_hoja_datos(
    ws: Worksheet, num_columnas: int, congelar_panel: bool = True
) -> None:
    """Aplica el estilo de encabezado azul/blanco institucional (fondo
    #0F3A7A, texto blanco en negrita) a la fila 1 de una hoja ya escrita con
    ``DataFrame.to_excel(writer, ..., sheet_name=...)`` — para unificar la
    identidad visual de las hojas del Informe completo con la del propio
    Diccionario de Datos."""
    for col_idx in range(1, num_columnas + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.fill = _FILL_ENCABEZADO
        celda.font = _FONT_ENCABEZADO
        celda.alignment = Alignment(wrap_text=True, vertical="center")
    if congelar_panel:
        ws.freeze_panes = ws.cell(row=2, column=1)


def ajustar_ancho_columnas_auto(
    ws: Worksheet, df: pd.DataFrame, ancho_min: int = 10, ancho_max: int = 45
) -> None:
    """Ajusta el ancho de cada columna de ``ws`` al contenido real de ``df``
    (encabezado + valores), acotado entre ``ancho_min`` y ``ancho_max``. El
    contenido de datos queda con ``wrap_text`` para seguir siendo legible sin
    una columna gigante."""
    for idx, columna in enumerate(df.columns, start=1):
        # pd.notna() como guardia: con el backend Arrow de pandas, astype(str)
        # no siempre convierte los nulos, y len(float) lanzaría TypeError.
        if len(df):
            max_len_datos = max(
                (len(str(v)) for v in df[columna] if pd.notna(v)), default=0
            )
        else:
            max_len_datos = 0
        ancho_encabezado = len(str(columna)) + 4
        ancho_datos = max(ancho_min, min(max_len_datos + 2, ancho_max))
        ws.column_dimensions[get_column_letter(idx)].width = max(
            ancho_encabezado, ancho_datos
        )

    ultima_fila = 1 + len(df)
    for fila in ws.iter_rows(min_row=2, max_row=ultima_fila, max_col=len(df.columns)):
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, vertical="top")


def escribir_hoja_descripcion_campos(
    writer: pd.ExcelWriter,
    hojas: dict[str, pd.DataFrame],
    sheet_name: str = "Descripción de campos",
) -> None:
    """Escribe la hoja "Diccionario de Datos" dentro del ``writer`` de
    pandas/openpyxl ya abierto, con un encabezado de identificación
    institucional (estilo "Ficha técnica" de los Lineamientos ONE) seguido de
    la tabla de metadatos de cada hoja del Informe completo.

    ``hojas`` es el mismo dict {nombre_de_hoja: DataFrame} que se exportó al
    archivo, para documentar exactamente las columnas presentes.
    """
    tabla = construir_tabla_descripcion_campos(hojas)

    FILA_INICIO_TABLA = 9  # espacio (0-indexed) para el bloque de ficha técnica
    tabla.to_excel(
        writer, index=False, sheet_name=sheet_name, startrow=FILA_INICIO_TABLA
    )

    ws: Worksheet = writer.sheets[sheet_name]

    # ── Bloque de identificación institucional ──────────────────────────
    hoy = _dt.date.today().strftime("%d/%m/%Y")
    ficha = [
        ("Nombre de la publicación", "Descripción de campos — Informe completo del Autodiagnóstico (CNBPEO)"),
        ("Institución", "Oficina Nacional de Estadística (ONE), República Dominicana"),
        (
            "Objetivo general",
            "Documentar la estructura, tipo de dato y significado de cada "
            "variable exportada en este archivo, conforme a los lineamientos "
            "de la ONE en la materia.",
        ),
        (
            "Cobertura",
            "Hojas 'Autodiagnóstico', 'Plan de acción', 'Estadísticas' y "
            "'Puntaje por nivel' de este mismo archivo, según la institución "
            "y los datos registrados al momento de generarlo.",
        ),
        ("Clasificación", "Conjunto de datos pasivo (archivo XLSX)"),
        ("Fecha de generación", hoy),
    ]

    ws.cell(row=1, column=1, value="Descripción de campos — CNBPEO").font = _FONT_TITULO
    fila = 2
    for etiqueta, valor in ficha:
        c_etq = ws.cell(row=fila, column=1, value=etiqueta)
        c_etq.font = _FONT_ETIQUETA_FICHA
        c_val = ws.cell(row=fila, column=2, value=valor)
        c_val.font = _FONT_NORMAL
        c_val.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=len(_COLUMNAS_TABLA))
        fila += 1

    # ── Formato de la tabla (encabezados en fila FILA_INICIO_TABLA + 1) ──
    fila_encabezado_tabla = FILA_INICIO_TABLA + 1
    for col_idx in range(1, len(_COLUMNAS_TABLA) + 1):
        celda = ws.cell(row=fila_encabezado_tabla, column=col_idx)
        celda.fill = _FILL_ENCABEZADO
        celda.font = _FONT_ENCABEZADO
        celda.alignment = Alignment(wrap_text=True, vertical="center")

    for col_idx, ancho in enumerate(_ANCHO_COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    for fila_idx in range(fila_encabezado_tabla + 1, fila_encabezado_tabla + 1 + len(tabla)):
        for col_idx in range(1, len(_COLUMNAS_TABLA) + 1):
            ws.cell(row=fila_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        ws.cell(row=fila_idx, column=1).font = Font(bold=True, name="Arial")

    ws.freeze_panes = ws.cell(row=fila_encabezado_tabla + 1, column=1)
