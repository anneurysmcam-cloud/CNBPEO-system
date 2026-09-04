"""
models/crud_autodiagnostico.py
===============================
CRUD y estadísticas del Autodiagnóstico de Calidad (Código Nacional de Buenas
Prácticas para las Estadísticas Oficiales, ONE — Matriz versión 004).

Jerarquía completa (igual que el Excel de referencia):

    Nivel (3: GEI/GPE/GRE)
      > Principio (15)
        > Requisito de cumplimiento (67 — tabla autodx_elementos)
          > Código de elemento / Verificación (252 — tabla autodx_verificaciones)

POR INSTITUCIÓN: cada institución pública del SEN llena y guarda sus propios
252 elementos por separado (tabla autodx_evaluaciones_inst, con clave única
institución+verificación). El plan de acción y el seguimiento también son por
institución. La lista base de instituciones está en
data/autodiagnostico_catalogo.py (INSTITUCIONES) y los usuarios pueden agregar
otras desde la app.

El cumplimiento se marca en el nivel de las 252 verificaciones ("Código de
elemento" del Excel, ej. GEI-1.1.1), exactamente como el Excel original.

Fórmula de % de cumplimiento (replica la hoja RESULTADOS del Excel):

    % por principio = (Iniciado + Parcial + Total) / (No Iniciado + Iniciado
                        + Parcial + Total) × 100          (excluye "No aplica")
    % por nivel = promedio simple de los % de sus principios
    % general    = promedio simple de los % de los 3 niveles

El esquema nuevo se crea y siembra de forma perezosa desde aquí mismo
(_asegurar_esquema_v2), así no hace falta tocar data/database.py.
"""

import datetime
import logging
import os
import re
import sqlite3
from io import BytesIO

import openpyxl
import streamlit as st

from data import database as db_mod

logger = logging.getLogger(__name__)

_CACHE_TTL_CATALOGO = 600
_CACHE_TTL_EVALUACIONES = 30

NIVELES_CUMPLIMIENTO: list[str] = [
    "No aplica", "No Iniciado", "Iniciado", "Cumplimiento Parcial", "Cumplimiento Total",
]
OPCIONES_CUMPLE: list[str] = ["SI", "NO", "N/A"]

# Puntaje como en la hoja RESULTADOS del Excel: solo cuenta "Cumplimiento Total"
# (lo 100% cumplido) sobre los elementos que aplican (se excluye "No aplica").
_CATEGORIAS_AVANCE = ("Cumplimiento Total",)
_CATEGORIAS_DENOMINADOR = ("No Iniciado", "Iniciado", "Cumplimiento Parcial", "Cumplimiento Total")

_PATRON_VERIFICACION = re.compile(r"^([A-Z]{3}-\d+\.\d+)\.\d+$")
_HOJAS_ESPERADAS = ("GEI", "GPE", "GRE")


def _invalidar_cache_evaluaciones() -> None:
    listar_verificaciones.clear()
    calcular_resumen.clear()


# ---------------------------------------------------------------------------
# Esquema v2 (perezoso) — verificaciones, evaluaciones por institución,
# instituciones, plan de acción y seguimiento (por institución).
# ---------------------------------------------------------------------------

_ESQUEMA_V2_LISTO = False


def _columna_existe(cursor, tabla: str, columna: str) -> bool:
    return any(r[1] == columna for r in cursor.execute(f"PRAGMA table_info({tabla})").fetchall())


def _crear_tablas_v2(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_verificaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            elemento_id INTEGER NOT NULL,
            codigo TEXT UNIQUE NOT NULL,
            texto TEXT,
            orden INTEGER NOT NULL,
            FOREIGN KEY (elemento_id) REFERENCES autodx_elementos(id) ON DELETE CASCADE
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_instituciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            orden INTEGER NOT NULL DEFAULT 50
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_evaluaciones_inst (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            institucion_id INTEGER NOT NULL,
            verificacion_id INTEGER NOT NULL,
            cumple TEXT CHECK(cumple IN ('SI','NO','N/A') OR cumple IS NULL),
            nivel_cumplimiento TEXT NOT NULL DEFAULT 'No Iniciado'
                CHECK(nivel_cumplimiento IN
                    ('No aplica','No Iniciado','Iniciado','Cumplimiento Parcial','Cumplimiento Total')),
            evidencia_actual TEXT,
            evidencia_anterior TEXT,
            comentario TEXT,
            accion_mejora TEXT,
            responsable TEXT,
            fecha_cumplimiento TEXT,
            actualizado_por INTEGER,
            fecha_actualizacion DATETIME DEFAULT (datetime('now')),
            UNIQUE(institucion_id, verificacion_id),
            FOREIGN KEY (institucion_id) REFERENCES autodx_instituciones(id) ON DELETE CASCADE,
            FOREIGN KEY (verificacion_id) REFERENCES autodx_verificaciones(id) ON DELETE CASCADE,
            FOREIGN KEY (actualizado_por) REFERENCES usuarios(id)
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_plan_accion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requerimiento TEXT, accion_mejora TEXT, actividades TEXT, insumo TEXT,
            presupuesto TEXT, fecha_cumplimiento TEXT, responsable TEXT,
            indicador_verificable TEXT, riesgo TEXT, acciones_mitigacion TEXT,
            observaciones TEXT, actualizado_por INTEGER,
            fecha_actualizacion DATETIME DEFAULT (datetime('now')),
            FOREIGN KEY (actualizado_por) REFERENCES usuarios(id)
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_seguimiento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL, descripcion TEXT, fecha_snapshot TEXT NOT NULL,
            general REAL, creado_por INTEGER,
            FOREIGN KEY (creado_por) REFERENCES usuarios(id)
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_seguimiento_detalle (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seguimiento_id INTEGER NOT NULL, nivel_codigo TEXT NOT NULL,
            principio_numero INTEGER NOT NULL, principio_nombre TEXT, score REAL,
            FOREIGN KEY (seguimiento_id) REFERENCES autodx_seguimiento(id) ON DELETE CASCADE
        )
        """
    )
    cursor.execute("CREATE TABLE IF NOT EXISTS autodx_config (clave TEXT PRIMARY KEY, valor TEXT)")
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS autodx_material (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            nombre_archivo TEXT NOT NULL,
            mime TEXT,
            contenido BLOB NOT NULL,
            subido_por INTEGER,
            fecha_subida TEXT,
            FOREIGN KEY (subido_por) REFERENCES usuarios(id)
        )
        """
    )

    # institucion_id en plan y seguimiento (por institución) — ALTER si falta
    if not _columna_existe(cursor, "autodx_plan_accion", "institucion_id"):
        cursor.execute("ALTER TABLE autodx_plan_accion ADD COLUMN institucion_id INTEGER")
    if not _columna_existe(cursor, "autodx_seguimiento", "institucion_id"):
        cursor.execute("ALTER TABLE autodx_seguimiento ADD COLUMN institucion_id INTEGER")
    if not _columna_existe(cursor, "autodx_material", "categoria"):
        cursor.execute("ALTER TABLE autodx_material ADD COLUMN categoria TEXT")

    cursor.execute("CREATE INDEX IF NOT EXISTS idx_autodx_verif_elemento ON autodx_verificaciones(elemento_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_autodx_eval_inst ON autodx_evaluaciones_inst(institucion_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_autodx_segdet_seg ON autodx_seguimiento_detalle(seguimiento_id)")


def _sembrar_verificaciones(cursor) -> None:
    from data.autodiagnostico_catalogo import VERIFICACIONES
    elemento_ids = {r[0]: r[1] for r in cursor.execute("SELECT codigo, id FROM autodx_elementos").fetchall()}
    for codigo, elemento_codigo, texto, orden in VERIFICACIONES:
        elemento_id = elemento_ids.get(elemento_codigo)
        if elemento_id is None:
            continue
        cursor.execute(
            """
            INSERT INTO autodx_verificaciones (elemento_id, codigo, texto, orden)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(codigo) DO UPDATE SET
                elemento_id = excluded.elemento_id, texto = excluded.texto, orden = excluded.orden
            """,
            (elemento_id, codigo, texto, orden),
        )


def _sembrar_instituciones(cursor) -> None:
    from data.autodiagnostico_catalogo import INSTITUCIONES
    for nombre, orden in INSTITUCIONES:
        cursor.execute(
            "INSERT INTO autodx_instituciones (nombre, orden) VALUES (?, ?) "
            "ON CONFLICT(nombre) DO UPDATE SET orden = excluded.orden",
            (nombre, orden),
        )


def _migrar_datos_a_institucion_defecto(cursor) -> None:
    """Si hay datos del modelo viejo (single-tenant) o filas de plan/seguimiento
    sin institución, se asignan a la institución por defecto (la ONE), para no
    perder lo que el usuario ya había cargado."""
    fila = cursor.execute("SELECT id FROM autodx_instituciones ORDER BY orden, id LIMIT 1").fetchone()
    if not fila:
        return
    inst_defecto = fila[0]

    # Migrar evaluaciones del modelo viejo (autodx_eval_verificaciones) si existe
    existe_vieja = cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='autodx_eval_verificaciones'"
    ).fetchone()
    if existe_vieja:
        hay_nuevas = cursor.execute(
            "SELECT COUNT(*) FROM autodx_evaluaciones_inst WHERE institucion_id = ?", (inst_defecto,)
        ).fetchone()[0]
        if hay_nuevas == 0:
            cursor.execute(
                """
                INSERT OR IGNORE INTO autodx_evaluaciones_inst
                    (institucion_id, verificacion_id, cumple, nivel_cumplimiento,
                     evidencia_actual, evidencia_anterior, comentario, accion_mejora,
                     responsable, fecha_cumplimiento, actualizado_por, fecha_actualizacion)
                SELECT ?, verificacion_id, cumple, nivel_cumplimiento, evidencia_actual,
                       evidencia_anterior, comentario, accion_mejora, responsable,
                       fecha_cumplimiento, actualizado_por, fecha_actualizacion
                FROM autodx_eval_verificaciones
                WHERE nivel_cumplimiento <> 'No Iniciado' OR cumple IS NOT NULL
                   OR evidencia_actual IS NOT NULL OR comentario IS NOT NULL
                """,
                (inst_defecto,),
            )

    # Plan y seguimiento sin institución -> institución por defecto
    cursor.execute("UPDATE autodx_plan_accion SET institucion_id = ? WHERE institucion_id IS NULL", (inst_defecto,))
    cursor.execute("UPDATE autodx_seguimiento SET institucion_id = ? WHERE institucion_id IS NULL", (inst_defecto,))


def _asegurar_esquema_v2() -> None:
    global _ESQUEMA_V2_LISTO
    if _ESQUEMA_V2_LISTO:
        return
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        _crear_tablas_v2(cursor)
        if cursor.execute("SELECT COUNT(*) FROM autodx_verificaciones").fetchone()[0] < 252:
            _sembrar_verificaciones(cursor)
        if cursor.execute("SELECT COUNT(*) FROM autodx_instituciones").fetchone()[0] == 0:
            _sembrar_instituciones(cursor)
        else:
            _sembrar_instituciones(cursor)  # idempotente: mantiene la lista base al día
        _migrar_datos_a_institucion_defecto(cursor)
        conn.commit()
        _ESQUEMA_V2_LISTO = True
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Instituciones
# ---------------------------------------------------------------------------

def listar_instituciones() -> list[dict]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        return [dict(r) for r in cursor.execute(
            "SELECT id, nombre, orden FROM autodx_instituciones ORDER BY orden, nombre"
        ).fetchall()]
    finally:
        conn.close()


def id_institucion_por_defecto() -> int | None:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        fila = cursor.execute("SELECT id FROM autodx_instituciones ORDER BY orden, id LIMIT 1").fetchone()
        return fila[0] if fila else None
    finally:
        conn.close()


def crear_o_obtener_institucion(nombre: str) -> tuple[int | None, str]:
    """Devuelve el id de la institución con ese nombre, creándola si no existe."""
    nombre = (nombre or "").strip()
    if not nombre:
        return None, "El nombre de la institución no puede estar vacío."
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        fila = cursor.execute("SELECT id FROM autodx_instituciones WHERE nombre = ?", (nombre,)).fetchone()
        if fila:
            return fila[0], "La institución ya existía."
        cursor.execute("INSERT INTO autodx_instituciones (nombre, orden) VALUES (?, 60)", (nombre,))
        conn.commit()
        return cursor.lastrowid, f"Institución '{nombre}' agregada."
    except sqlite3.Error as exc:
        conn.rollback()
        return None, f"No se pudo agregar la institución: {exc}"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Catálogo (solo lectura): niveles y principios
# ---------------------------------------------------------------------------

@st.cache_data(ttl=_CACHE_TTL_CATALOGO)
def listar_niveles() -> list[dict]:
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        return [dict(f) for f in cursor.execute("SELECT * FROM autodx_niveles ORDER BY orden").fetchall()]
    finally:
        conn.close()


@st.cache_data(ttl=_CACHE_TTL_CATALOGO)
def listar_principios(nivel_codigo: str | None = None) -> list[dict]:
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        sql = """
            SELECT p.id, p.numero, p.nombre, p.orden,
                   n.id AS nivel_id, n.codigo AS nivel_codigo, n.nombre AS nivel_nombre
            FROM autodx_principios p JOIN autodx_niveles n ON n.id = p.nivel_id
        """
        params: list = []
        if nivel_codigo:
            sql += " WHERE n.codigo = ?"
            params.append(nivel_codigo)
        sql += " ORDER BY p.orden"
        return [dict(f) for f in cursor.execute(sql, params).fetchall()]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Verificaciones (252) por institución — lectura y marcado
# ---------------------------------------------------------------------------

@st.cache_data(ttl=_CACHE_TTL_EVALUACIONES)
def listar_verificaciones(
    institucion_id: int,
    nivel_codigo: str | None = None,
    principio_numero: int | None = None,
    requisito_codigo: str | None = None,
) -> list[dict]:
    """Las 252 verificaciones con la evaluación de la institución indicada."""
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        sql = """
            SELECT
                v.id AS verificacion_id, v.codigo, v.texto, v.orden,
                e.id AS elemento_id, e.codigo AS requisito_codigo, e.requisito_cumplimiento,
                p.numero AS principio_numero, p.nombre AS principio_nombre,
                n.codigo AS nivel_codigo, n.nombre AS nivel_nombre,
                ev.cumple, ev.nivel_cumplimiento, ev.evidencia_actual,
                ev.evidencia_anterior, ev.comentario, ev.accion_mejora,
                ev.responsable, ev.fecha_cumplimiento, ev.fecha_actualizacion
            FROM autodx_verificaciones v
            JOIN autodx_elementos e ON e.id = v.elemento_id
            JOIN autodx_principios p ON p.id = e.principio_id
            JOIN autodx_niveles n ON n.id = p.nivel_id
            LEFT JOIN autodx_evaluaciones_inst ev
                   ON ev.verificacion_id = v.id AND ev.institucion_id = ?
            WHERE 1=1
        """
        params: list = [institucion_id]
        if nivel_codigo:
            sql += " AND n.codigo = ?"
            params.append(nivel_codigo)
        if principio_numero is not None:
            sql += " AND p.numero = ?"
            params.append(principio_numero)
        if requisito_codigo:
            sql += " AND e.codigo = ?"
            params.append(requisito_codigo)
        sql += " ORDER BY v.orden"
        return [dict(f) for f in cursor.execute(sql, params).fetchall()]
    finally:
        conn.close()


def listar_requisitos(nivel_codigo: str | None = None, principio_numero: int | None = None) -> list[dict]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        sql = """
            SELECT e.id, e.codigo, e.numero, e.requisito_cumplimiento,
                   p.numero AS principio_numero, n.codigo AS nivel_codigo
            FROM autodx_elementos e
            JOIN autodx_principios p ON p.id = e.principio_id
            JOIN autodx_niveles n ON n.id = p.nivel_id
            WHERE 1=1
        """
        params: list = []
        if nivel_codigo:
            sql += " AND n.codigo = ?"
            params.append(nivel_codigo)
        if principio_numero is not None:
            sql += " AND p.numero = ?"
            params.append(principio_numero)
        sql += " ORDER BY e.orden"
        return [dict(f) for f in cursor.execute(sql, params).fetchall()]
    finally:
        conn.close()


def guardar_evaluacion_verificacion(
    institucion_id: int,
    verificacion_id: int,
    cumple: str | None,
    nivel_cumplimiento: str,
    evidencia_actual: str = "",
    evidencia_anterior: str = "",
    comentario: str = "",
    accion_mejora: str = "",
    responsable: str = "",
    fecha_cumplimiento: str | None = None,
    usuario_id: int | None = None,
) -> tuple[bool, str]:
    if cumple not in (None, "", *OPCIONES_CUMPLE):
        return False, f"'¿Cumple?' debe ser uno de {OPCIONES_CUMPLE} o quedar sin marcar."
    if nivel_cumplimiento not in NIVELES_CUMPLIMIENTO:
        return False, f"Nivel de cumplimiento debe ser uno de {NIVELES_CUMPLIMIENTO}."
    cumple = cumple or None

    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        fila = cursor.execute("SELECT codigo FROM autodx_verificaciones WHERE id = ?", (verificacion_id,)).fetchone()
        if not fila:
            return False, "La verificación indicada no existe."
        codigo = fila[0]
        cursor.execute(
            """
            INSERT INTO autodx_evaluaciones_inst
                (institucion_id, verificacion_id, cumple, nivel_cumplimiento, evidencia_actual,
                 evidencia_anterior, comentario, accion_mejora, responsable,
                 fecha_cumplimiento, actualizado_por, fecha_actualizacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(institucion_id, verificacion_id) DO UPDATE SET
                cumple = excluded.cumple, nivel_cumplimiento = excluded.nivel_cumplimiento,
                evidencia_actual = excluded.evidencia_actual, evidencia_anterior = excluded.evidencia_anterior,
                comentario = excluded.comentario, accion_mejora = excluded.accion_mejora,
                responsable = excluded.responsable, fecha_cumplimiento = excluded.fecha_cumplimiento,
                actualizado_por = excluded.actualizado_por, fecha_actualizacion = datetime('now')
            """,
            (
                institucion_id, verificacion_id, cumple, nivel_cumplimiento, evidencia_actual or None,
                evidencia_anterior or None, comentario or None, accion_mejora or None,
                responsable or None, fecha_cumplimiento or None, usuario_id,
            ),
        )
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_EVALUAR",
                      f"Inst {institucion_id} · Verificación '{codigo}': cumple={cumple}, nivel={nivel_cumplimiento}.")
        conn.commit()
        _invalidar_cache_evaluaciones()
        return True, f"Cumplimiento de '{codigo}' guardado correctamente."
    except sqlite3.Error as exc:
        conn.rollback()
        logger.error("Error al guardar evaluación (inst=%s, verif=%s): %s", institucion_id, verificacion_id, exc)
        return False, f"No se pudo guardar: {exc}"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Estadísticas (% de cumplimiento sobre las 252, por institución)
# ---------------------------------------------------------------------------

def _score(conteos: dict[str, int]) -> float | None:
    denominador = sum(conteos.get(c, 0) for c in _CATEGORIAS_DENOMINADOR)
    if denominador == 0:
        return None
    numerador = sum(conteos.get(c, 0) for c in _CATEGORIAS_AVANCE)
    return round(numerador / denominador * 100, 1)


@st.cache_data(ttl=_CACHE_TTL_EVALUACIONES)
def calcular_resumen(institucion_id: int) -> dict:
    verificaciones = listar_verificaciones(institucion_id)
    _ETIQUETA_A_CLAVE = {
        "No aplica": "no_aplica", "No Iniciado": "no_iniciado", "Iniciado": "iniciado",
        "Cumplimiento Parcial": "parcial", "Cumplimiento Total": "total",
    }
    por_principio: dict[tuple, dict] = {}
    for v in verificaciones:
        clave = (v["nivel_codigo"], v["principio_numero"])
        if clave not in por_principio:
            por_principio[clave] = {
                "nivel_codigo": v["nivel_codigo"], "nivel_nombre": v["nivel_nombre"],
                "principio_numero": v["principio_numero"], "principio_nombre": v["principio_nombre"],
                "no_aplica": 0, "no_iniciado": 0, "iniciado": 0, "parcial": 0, "total": 0,
            }
        nivel_cump = v["nivel_cumplimiento"] or "No Iniciado"
        por_principio[clave][_ETIQUETA_A_CLAVE[nivel_cump]] += 1

    resultado_principios = []
    for datos in por_principio.values():
        conteos_formula = {
            "No Iniciado": datos["no_iniciado"], "Iniciado": datos["iniciado"],
            "Cumplimiento Parcial": datos["parcial"], "Cumplimiento Total": datos["total"],
        }
        total_elementos = sum(datos[c] for c in ("no_aplica", "no_iniciado", "iniciado", "parcial", "total"))
        resultado_principios.append({**datos, "total_elementos": total_elementos, "score": _score(conteos_formula)})
    resultado_principios.sort(key=lambda d: (d["nivel_codigo"], d["principio_numero"]))

    por_nivel: dict[str, dict] = {}
    for p in resultado_principios:
        nc = p["nivel_codigo"]
        por_nivel.setdefault(nc, {"nivel_codigo": nc, "nivel_nombre": p["nivel_nombre"], "_scores": []})
        if p["score"] is not None:
            por_nivel[nc]["_scores"].append(p["score"])
    resultado_niveles = []
    for datos in por_nivel.values():
        scores = datos.pop("_scores")
        datos["score"] = round(sum(scores) / len(scores), 1) if scores else None
        resultado_niveles.append(datos)
    resultado_niveles.sort(key=lambda d: d["nivel_codigo"])

    scores_nivel = [n["score"] for n in resultado_niveles if n["score"] is not None]
    general = round(sum(scores_nivel) / len(scores_nivel), 1) if scores_nivel else None
    return {"por_principio": resultado_principios, "por_nivel": resultado_niveles, "general": general}


# ---------------------------------------------------------------------------
# Importación desde el Excel original (a nivel de las 252, por institución)
# ---------------------------------------------------------------------------

def _texto(valor) -> str | None:
    if valor is None:
        return None
    t = str(valor).strip()
    return t or None


def _normalizar_cumple(valor: str | None) -> str | None:
    if not valor:
        return None
    v = valor.strip().upper().replace(" ", "")
    if v in ("SI", "SÍ", "S"):
        return "SI"
    if v == "NO":
        return "NO"
    if v in ("N/A", "NA", "NOAPLICA"):
        return "N/A"
    return None


def _normalizar_nivel(valor: str | None) -> str | None:
    if not valor:
        return None
    mapa = {
        "no aplica": "No aplica", "no iniciado": "No Iniciado", "iniciado": "Iniciado",
        "cumplimiento parcial": "Cumplimiento Parcial", "parcial": "Cumplimiento Parcial",
        "cumplimiento total": "Cumplimiento Total", "total": "Cumplimiento Total",
    }
    return mapa.get(valor.strip().lower())


def _fecha_iso(valor) -> str | None:
    if valor is None:
        return None
    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat()
    if isinstance(valor, datetime.date):
        return valor.isoformat()
    try:
        return datetime.date.fromisoformat(str(valor).strip()[:10]).isoformat()
    except ValueError:
        return None


def _norm_encabezado_excel(valor) -> str:
    import unicodedata
    s = unicodedata.normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode("ascii")
    s = s.lower().replace("¿", "").replace("?", "").strip()
    return " ".join(s.split())


# Nombre de encabezado (normalizado) -> campo interno.
_ENCABEZADOS_EXCEL = {
    "codigo de elemento": "codigo", "codigo del elemento": "codigo", "codigo": "codigo",
    "cumple con el elemento": "cumple", "cumple": "cumple",
    "nivel de cumplimiento": "nivel", "nivel": "nivel",
    "evidencia actual": "evidencia_actual", "evidencia": "evidencia_actual",
    "evidencia anterior": "evidencia_anterior",
    "comentario": "comentario", "comentarios": "comentario",
    "accion de mejora": "accion_mejora", "acciones de mejora": "accion_mejora",
    "responsable": "responsable",
    "fecha de cumplimiento": "fecha_cumplimiento", "fecha cumplimiento": "fecha_cumplimiento",
    "fecha": "fecha_cumplimiento",
}
# Posiciones por defecto (respaldo si no se encuentran los encabezados).
_COLS_POR_DEFECTO = {"codigo": 3, "cumple": 5, "nivel": 6, "evidencia_actual": 7,
                     "evidencia_anterior": 8, "comentario": 9, "accion_mejora": 10,
                     "responsable": 11, "fecha_cumplimiento": 12}


def _leer_verificaciones_excel(contenido: bytes) -> tuple[dict[str, dict] | None, str | None]:
    try:
        libro = openpyxl.load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:
        return None, f"No se pudo leer el archivo — ¿es un .xlsx válido? ({exc})"
    faltantes = [h for h in _HOJAS_ESPERADAS if h not in libro.sheetnames]
    if faltantes:
        libro.close()
        return None, (
            f"El archivo no tiene las hojas esperadas ({', '.join(faltantes)} no encontrada(s)). "
            "Debe ser la Matriz de Autodiagnóstico original, con hojas GEI, GPE y GRE."
        )

    filas: dict[str, dict] = {}
    for nombre_hoja in _HOJAS_ESPERADAS:
        hoja = libro[nombre_hoja]
        todas = list(hoja.iter_rows(values_only=True))
        # Ubicar columnas por NOMBRE de encabezado (robusto a corrimientos de columnas).
        colmap, inicio = None, 0
        for r, fila in enumerate(todas):
            encabezados = {}
            for idx, val in enumerate(fila):
                campo = _ENCABEZADOS_EXCEL.get(_norm_encabezado_excel(val))
                if campo and campo not in encabezados:
                    encabezados[campo] = idx
            if "codigo" in encabezados and "cumple" in encabezados and "nivel" in encabezados:
                colmap, inicio = encabezados, r + 1
                break
        if colmap is None:
            colmap, inicio = dict(_COLS_POR_DEFECTO), 0  # respaldo a posiciones fijas

        def _celda(fila, campo):
            j = colmap.get(campo)
            return fila[j] if j is not None and len(fila) > j else None

        for fila in todas[inicio:]:
            codigo = _celda(fila, "codigo")
            if not codigo or not _PATRON_VERIFICACION.match(str(codigo).strip()):
                continue
            cod = str(codigo).strip()
            filas[cod] = {
                "cumple": _normalizar_cumple(_texto(_celda(fila, "cumple"))),
                "nivel": _normalizar_nivel(_texto(_celda(fila, "nivel"))),
                "evidencia_actual": _texto(_celda(fila, "evidencia_actual")),
                "evidencia_anterior": _texto(_celda(fila, "evidencia_anterior")),
                "comentario": _texto(_celda(fila, "comentario")),
                "accion_mejora": _texto(_celda(fila, "accion_mejora")),
                "responsable": _texto(_celda(fila, "responsable")),
                "fecha_cumplimiento": _fecha_iso(_celda(fila, "fecha_cumplimiento")),
            }
    libro.close()
    if not filas:
        return None, ("No se encontraron filas con un 'Código de elemento' reconocible "
                      "(ej. GEI-1.1.1) en las hojas GEI/GPE/GRE.")
    return filas, None


_CAMPOS_VERIF = (
    "cumple", "nivel", "evidencia_actual", "evidencia_anterior",
    "comentario", "accion_mejora", "responsable", "fecha_cumplimiento",
)


def previsualizar_importacion(contenido: bytes) -> tuple[bool, str, dict]:
    filas, error = _leer_verificaciones_excel(contenido)
    if error:
        return False, error, {}
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        catalogo = {r[0] for r in cursor.execute("SELECT codigo FROM autodx_verificaciones").fetchall()}
    finally:
        conn.close()
    con_datos, sin_datos, no_reconocidos = [], 0, []
    for cod, datos in filas.items():
        if cod not in catalogo:
            no_reconocidos.append(cod)
            continue
        if any(datos[c] for c in _CAMPOS_VERIF):
            con_datos.append({"codigo": cod, **datos})
        else:
            sin_datos += 1
    resumen = {"total_en_archivo": len(filas), "con_datos": con_datos,
               "sin_datos": sin_datos, "no_reconocidos": sorted(no_reconocidos)}
    mensaje = (f"El archivo trae información para {len(con_datos)} de las 252 verificaciones "
               f"({sin_datos} vienen vacías en el archivo).")
    return True, mensaje, resumen


def importar_desde_excel(
    institucion_id: int, contenido: bytes, nombre_archivo: str, usuario_id: int | None
) -> tuple[bool, str, dict]:
    filas, error = _leer_verificaciones_excel(contenido)
    if error:
        return False, error, {}
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        catalogo = {r["codigo"]: r["id"] for r in cursor.execute("SELECT id, codigo FROM autodx_verificaciones").fetchall()}
        existentes = {
            r["verificacion_id"]: dict(r)
            for r in cursor.execute(
                "SELECT verificacion_id, cumple, nivel_cumplimiento, evidencia_actual, "
                "evidencia_anterior, comentario, accion_mejora, responsable, fecha_cumplimiento "
                "FROM autodx_evaluaciones_inst WHERE institucion_id = ?", (institucion_id,)
            ).fetchall()
        }
        actualizados, sin_datos, no_reconocidos = 0, 0, []
        for cod, datos in filas.items():
            verif_id = catalogo.get(cod)
            if verif_id is None:
                no_reconocidos.append(cod)
                continue
            if not any(datos[c] for c in _CAMPOS_VERIF):
                sin_datos += 1
                continue
            prev = existentes.get(verif_id, {})
            cumple = datos["cumple"] or prev.get("cumple")
            nivel = datos["nivel"] or prev.get("nivel_cumplimiento") or "No Iniciado"
            evidencia_actual = datos["evidencia_actual"] or prev.get("evidencia_actual")
            evidencia_anterior = datos["evidencia_anterior"] or prev.get("evidencia_anterior")
            comentario = datos["comentario"] or prev.get("comentario")
            accion_mejora = datos["accion_mejora"] or prev.get("accion_mejora")
            responsable = datos["responsable"] or prev.get("responsable")
            fecha_cumplimiento = datos["fecha_cumplimiento"] or prev.get("fecha_cumplimiento")
            cursor.execute(
                """
                INSERT INTO autodx_evaluaciones_inst
                    (institucion_id, verificacion_id, cumple, nivel_cumplimiento, evidencia_actual,
                     evidencia_anterior, comentario, accion_mejora, responsable,
                     fecha_cumplimiento, actualizado_por, fecha_actualizacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(institucion_id, verificacion_id) DO UPDATE SET
                    cumple = excluded.cumple, nivel_cumplimiento = excluded.nivel_cumplimiento,
                    evidencia_actual = excluded.evidencia_actual, evidencia_anterior = excluded.evidencia_anterior,
                    comentario = excluded.comentario, accion_mejora = excluded.accion_mejora,
                    responsable = excluded.responsable, fecha_cumplimiento = excluded.fecha_cumplimiento,
                    actualizado_por = excluded.actualizado_por, fecha_actualizacion = datetime('now')
                """,
                (institucion_id, verif_id, cumple, nivel, evidencia_actual, evidencia_anterior,
                 comentario, accion_mejora, responsable, fecha_cumplimiento, usuario_id),
            )
            actualizados += 1

        filas_plan = _importar_plan_desde_workbook(contenido, cursor, institucion_id, usuario_id)

        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_IMPORTAR_EXCEL",
                      f"Inst {institucion_id} · Archivo '{nombre_archivo}': {actualizados} verificaciones, "
                      f"{sin_datos} sin datos, {len(no_reconocidos)} no reconocidas, {filas_plan} plan.")
        conn.commit()
        _invalidar_cache_evaluaciones()
        resumen = {"actualizados": actualizados, "sin_datos": sin_datos,
                   "no_reconocidos": sorted(no_reconocidos), "filas_plan": filas_plan}
        mensaje = f"Importación completa: {actualizados} de las 252 verificaciones actualizadas."
        if filas_plan:
            mensaje += f" {filas_plan} fila(s) de plan de acción importada(s)."
        if no_reconocidos:
            mensaje += f" ⚠️ {len(no_reconocidos)} código(s) no reconocido(s)."
        return True, mensaje, resumen
    except sqlite3.Error as exc:
        conn.rollback()
        logger.error("Error de BD al importar Excel de autodiagnóstico: %s", exc)
        return False, f"No se pudo importar: {exc}", {}
    finally:
        conn.close()


def _importar_plan_desde_workbook(contenido: bytes, cursor, institucion_id: int, usuario_id: int | None) -> int:
    try:
        libro = openpyxl.load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception:
        return 0
    # Buscar la hoja del plan sin importar tildes/mayúsculas ("Plan de acción" / "Plan de accion").
    nombre_plan = None
    for s in libro.sheetnames:
        if _norm_encabezado_excel(s) == "plan de accion":
            nombre_plan = s
            break
    if nombre_plan is None:
        libro.close()
        return 0
    hoja = libro[nombre_plan]
    existentes = {
        (r[0] or "", r[1] or "")
        for r in cursor.execute(
            "SELECT requerimiento, accion_mejora FROM autodx_plan_accion WHERE institucion_id = ?",
            (institucion_id,)
        ).fetchall()
    }
    filas_plan = list(hoja.iter_rows(values_only=True))
    # Detectar la fila de encabezado (la que contiene "Requerimiento") y empezar después.
    inicio = 8
    for r, fila in enumerate(filas_plan):
        if any(_norm_encabezado_excel(v) == "requerimiento" for v in fila if v is not None):
            inicio = r + 1
            break
    # Agrupar: cada requerimiento es UNA fila. Las sub-filas (con requerimiento
    # vacío, por la celda combinada del Excel) suman sus valores al requerimiento
    # de arriba, uniéndolos con saltos de línea.
    _CAMPOS_IDX = {"requerimiento": 1, "accion_mejora": 2, "actividades": 3, "insumo": 4,
                   "presupuesto": 5, "fecha_cumplimiento": 6, "responsable": 7,
                   "indicador_verificable": 8, "riesgo": 9, "acciones_mitigacion": 10,
                   "observaciones": 11}

    def _val(fila, idx):
        return (_texto(fila[idx]) if len(fila) > idx else None)

    grupos: list[dict] = []
    actual: dict | None = None
    for fila in filas_plan[inicio:]:
        req = _val(fila, 1)
        campos = {k: _val(fila, i) for k, i in _CAMPOS_IDX.items()}
        if not any(campos.values()):
            continue
        if req:
            actual = {k: (v or "") for k, v in campos.items()}
            grupos.append(actual)
        elif actual is not None:
            for k, v in campos.items():
                if v:
                    actual[k] = (actual[k] + "\n" + v) if actual.get(k) else v
        # (si no hay 'actual' todavía y el requerimiento está vacío, se ignora)

    insertadas = 0
    for g in grupos:
        clave = (g.get("requerimiento") or "", g.get("accion_mejora") or "")
        if clave in existentes:
            continue
        cursor.execute(
            """
            INSERT INTO autodx_plan_accion
                (institucion_id, requerimiento, accion_mejora, actividades, insumo, presupuesto,
                 fecha_cumplimiento, responsable, indicador_verificable, riesgo,
                 acciones_mitigacion, observaciones, actualizado_por, fecha_actualizacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (
                institucion_id, g.get("requerimiento") or None, g.get("accion_mejora") or None,
                g.get("actividades") or None, g.get("insumo") or None, g.get("presupuesto") or None,
                g.get("fecha_cumplimiento") or None, g.get("responsable") or None,
                g.get("indicador_verificable") or None, g.get("riesgo") or None,
                g.get("acciones_mitigacion") or None, g.get("observaciones") or None,
                usuario_id,
            ),
        )
        existentes.add(clave)
        insertadas += 1
    libro.close()
    return insertadas


# ---------------------------------------------------------------------------
# Plan de acción (por institución)
# ---------------------------------------------------------------------------

_CAMPOS_PLAN = (
    "requerimiento", "accion_mejora", "actividades", "insumo", "presupuesto",
    "fecha_cumplimiento", "responsable", "indicador_verificable", "riesgo",
    "acciones_mitigacion", "observaciones",
)


def listar_plan_accion(institucion_id: int) -> list[dict]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        return [dict(r) for r in cursor.execute(
            "SELECT * FROM autodx_plan_accion WHERE institucion_id = ? ORDER BY id", (institucion_id,)
        ).fetchall()]
    finally:
        conn.close()


def guardar_fila_plan(institucion_id: int, datos: dict, fila_id: int | None, usuario_id: int | None) -> tuple[bool, str]:
    _asegurar_esquema_v2()
    valores = {c: (datos.get(c) or None) for c in _CAMPOS_PLAN}
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        if fila_id is None:
            cols = ", ".join(_CAMPOS_PLAN)
            marcas = ", ".join(["?"] * len(_CAMPOS_PLAN))
            cursor.execute(
                f"INSERT INTO autodx_plan_accion (institucion_id, {cols}, actualizado_por, fecha_actualizacion) "
                f"VALUES (?, {marcas}, ?, datetime('now'))",
                (institucion_id, *[valores[c] for c in _CAMPOS_PLAN], usuario_id),
            )
            accion, msg = "AUTODX_PLAN_CREAR", "Fila de plan de acción agregada."
        else:
            asignaciones = ", ".join(f"{c} = ?" for c in _CAMPOS_PLAN)
            cursor.execute(
                f"UPDATE autodx_plan_accion SET {asignaciones}, actualizado_por = ?, "
                f"fecha_actualizacion = datetime('now') WHERE id = ? AND institucion_id = ?",
                (*[valores[c] for c in _CAMPOS_PLAN], usuario_id, fila_id, institucion_id),
            )
            accion, msg = "AUTODX_PLAN_EDITAR", "Fila de plan de acción actualizada."
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, accion, f"Inst {institucion_id} · Plan (id={fila_id}).")
        conn.commit()
        return True, msg
    except sqlite3.Error as exc:
        conn.rollback()
        return False, f"No se pudo guardar: {exc}"
    finally:
        conn.close()


def eliminar_fila_plan(institucion_id: int, fila_id: int, usuario_id: int | None) -> tuple[bool, str]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM autodx_plan_accion WHERE id = ? AND institucion_id = ?", (fila_id, institucion_id))
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_PLAN_ELIMINAR", f"Inst {institucion_id} · Plan (id={fila_id}).")
        conn.commit()
        return True, "Fila eliminada."
    except sqlite3.Error as exc:
        conn.rollback()
        return False, f"No se pudo eliminar: {exc}"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Seguimiento (snapshots por institución)
# ---------------------------------------------------------------------------

def guardar_seguimiento(institucion_id: int, nombre: str, descripcion: str, usuario_id: int | None) -> tuple[bool, str]:
    nombre = (nombre or "").strip()
    if not nombre:
        return False, "Ponle un nombre al seguimiento (ej. '1er Resultado')."
    _asegurar_esquema_v2()
    resumen = calcular_resumen(institucion_id)
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        fecha = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
        cursor.execute(
            "INSERT INTO autodx_seguimiento (institucion_id, nombre, descripcion, fecha_snapshot, general, creado_por) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (institucion_id, nombre, (descripcion or "").strip() or None, fecha, resumen["general"], usuario_id),
        )
        seg_id = cursor.lastrowid
        for p in resumen["por_principio"]:
            cursor.execute(
                "INSERT INTO autodx_seguimiento_detalle "
                "(seguimiento_id, nivel_codigo, principio_numero, principio_nombre, score) VALUES (?, ?, ?, ?, ?)",
                (seg_id, p["nivel_codigo"], p["principio_numero"], p["principio_nombre"], p["score"]),
            )
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_SEGUIMIENTO_GUARDAR",
                      f"Inst {institucion_id} · Seguimiento '{nombre}' (general={resumen['general']}).")
        conn.commit()
        return True, f"Seguimiento '{nombre}' guardado."
    except sqlite3.Error as exc:
        conn.rollback()
        return False, f"No se pudo guardar el seguimiento: {exc}"
    finally:
        conn.close()


def listar_seguimientos(institucion_id: int) -> list[dict]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        return [dict(r) for r in cursor.execute(
            "SELECT * FROM autodx_seguimiento WHERE institucion_id = ? ORDER BY id", (institucion_id,)
        ).fetchall()]
    finally:
        conn.close()


def obtener_detalle_seguimientos(institucion_id: int) -> list[dict]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        return [dict(r) for r in cursor.execute(
            "SELECT d.seguimiento_id, s.nombre, d.nivel_codigo, d.principio_numero, d.principio_nombre, d.score "
            "FROM autodx_seguimiento_detalle d JOIN autodx_seguimiento s ON s.id = d.seguimiento_id "
            "WHERE s.institucion_id = ? ORDER BY s.id, d.principio_numero", (institucion_id,)
        ).fetchall()]
    finally:
        conn.close()


def eliminar_seguimiento(institucion_id: int, seg_id: int, usuario_id: int | None) -> tuple[bool, str]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM autodx_seguimiento_detalle WHERE seguimiento_id = ?", (seg_id,))
        cursor.execute("DELETE FROM autodx_seguimiento WHERE id = ? AND institucion_id = ?", (seg_id, institucion_id))
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_SEGUIMIENTO_ELIMINAR", f"Inst {institucion_id} · Seguimiento id={seg_id}.")
        conn.commit()
        return True, "Seguimiento eliminado."
    except sqlite3.Error as exc:
        conn.rollback()
        return False, f"No se pudo eliminar: {exc}"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Configuración y sincronización automática desde un Excel fijo (por institución)
# ---------------------------------------------------------------------------

def _clave_ruta(institucion_id: int) -> str:
    return f"ruta_excel_autodx_{institucion_id}"


def _clave_mtime(institucion_id: int) -> str:
    return f"ultimo_mtime_excel_{institucion_id}"


def _clave_sync(institucion_id: int) -> str:
    return f"ultima_sync_excel_{institucion_id}"


def obtener_config(clave: str) -> str | None:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        fila = cursor.execute("SELECT valor FROM autodx_config WHERE clave = ?", (clave,)).fetchone()
        return fila[0] if fila else None
    finally:
        conn.close()


def guardar_config(clave: str, valor: str | None) -> None:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO autodx_config (clave, valor) VALUES (?, ?) "
            "ON CONFLICT(clave) DO UPDATE SET valor = excluded.valor",
            (clave, valor),
        )
        conn.commit()
    finally:
        conn.close()


def obtener_ruta_excel(institucion_id: int) -> str | None:
    return obtener_config(_clave_ruta(institucion_id))


def guardar_ruta_excel(institucion_id: int, ruta: str | None) -> None:
    guardar_config(_clave_ruta(institucion_id), (ruta or "").strip() or None)
    guardar_config(_clave_mtime(institucion_id), None)


def estado_sincronizacion(institucion_id: int) -> dict:
    ruta = obtener_ruta_excel(institucion_id)
    return {
        "ruta": ruta,
        "existe": bool(ruta) and os.path.isfile(ruta),
        "ultima_sync": obtener_config(_clave_sync(institucion_id)),
    }


def importar_desde_ruta(institucion_id: int, ruta: str | None, usuario_id: int | None) -> tuple[bool, str, dict]:
    if not ruta:
        return False, "No hay una ruta de Excel configurada.", {}
    if not os.path.isfile(ruta):
        return False, f"No se encontró ningún archivo en la ruta: {ruta}", {}
    try:
        with open(ruta, "rb") as fh:
            contenido = fh.read()
    except OSError as exc:
        return False, f"No se pudo leer el archivo: {exc}", {}
    ok, mensaje, resumen = importar_desde_excel(institucion_id, contenido, os.path.basename(ruta), usuario_id)
    if ok:
        try:
            guardar_config(_clave_mtime(institucion_id), str(os.path.getmtime(ruta)))
            guardar_config(_clave_sync(institucion_id), datetime.datetime.now().isoformat(sep=" ", timespec="seconds"))
        except OSError:
            pass
    return ok, mensaje, resumen


def sincronizar_si_cambio(institucion_id: int, usuario_id: int | None) -> tuple[bool, str, dict] | None:
    ruta = obtener_ruta_excel(institucion_id)
    if not ruta or not os.path.isfile(ruta):
        return None
    try:
        mtime_actual = str(os.path.getmtime(ruta))
    except OSError:
        return None
    if obtener_config(_clave_mtime(institucion_id)) == mtime_actual:
        return None
    return importar_desde_ruta(institucion_id, ruta, usuario_id)


# ---------------------------------------------------------------------------
# Exportación de resultados (Excel y PDF, por institución)
# ---------------------------------------------------------------------------

_LETRA_NIVEL_EXPORT = {"GEI": "A", "GPE": "B", "GRE": "C"}


def _filas_resultados(institucion_id: int) -> tuple[list[dict], list[dict], float | None]:
    resumen = calcular_resumen(institucion_id)
    filas = []
    puntaje_total_por_nivel: dict = {}
    for p in resumen["por_principio"]:
        total, parcial = p["total"], p["parcial"]
        iniciado, no_ini, no_aplica = p["iniciado"], p["no_iniciado"], p["no_aplica"]
        k = total + parcial + iniciado + no_ini
        pct = (lambda x: round(x / k * 100, 1) if k else None)
        pt_total = pct(total)
        filas.append({
            "nivel": _LETRA_NIVEL_EXPORT.get(p["nivel_codigo"], p["nivel_codigo"]),
            "numero": p["principio_numero"], "principio": p["principio_nombre"],
            "total": total, "parcial": parcial, "iniciado": iniciado,
            "no_iniciado": no_ini, "faltantes": no_aplica, "total_pautas": k,
            "pt_total": pt_total, "pt_parcial": pct(parcial),
            "pt_iniciado": pct(iniciado), "pt_no_iniciado": pct(no_ini),
            "total_pct": (round((pt_total or 0) + (pct(parcial) or 0) + (pct(iniciado) or 0) + (pct(no_ini) or 0), 1) if k else None),
        })
        if k:
            puntaje_total_por_nivel.setdefault(p["nivel_codigo"], []).append(pt_total)
    filas_nivel = []
    for nivel in resumen["por_nivel"]:
        vals = puntaje_total_por_nivel.get(nivel["nivel_codigo"], [])
        filas_nivel.append({
            "nivel": _LETRA_NIVEL_EXPORT.get(nivel["nivel_codigo"], nivel["nivel_codigo"]),
            "nombre": nivel["nivel_nombre"],
            "puntaje": round(sum(vals) / len(vals), 1) if vals else None,
        })
    return filas, filas_nivel, resumen["general"]


def _nombre_institucion(institucion_id: int) -> str:
    for i in listar_instituciones():
        if i["id"] == institucion_id:
            return i["nombre"]
    return "—"


def exportar_resultados_excel(institucion_id: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    filas, filas_nivel, general = _filas_resultados(institucion_id)
    nombre_inst = _nombre_institucion(institucion_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "RESULTADOS"
    negrita = Font(bold=True)
    ws.append(["Autodiagnóstico CNBPEO — Resultados desagregados por principio"])
    ws["A1"].font = negrita
    ws.append([f"Institución: {nombre_inst}"])
    ws["A2"].font = negrita
    ws.append([])
    encabezados = ["Nivel", "N°", "Principio", "Total", "Parcial", "Iniciado", "No Iniciado",
                   "Faltantes", "Total pautas", "Puntaje C. total", "Puntaje C. Parcial",
                   "Puntaje Iniciado", "Puntaje No iniciado", "Total %"]
    ws.append(encabezados)
    for c in ws[ws.max_row]:
        c.font = negrita
    for f in filas:
        ws.append([f["nivel"], f["numero"], f["principio"], f["total"], f["parcial"], f["iniciado"],
                   f["no_iniciado"], f["faltantes"], f["total_pautas"], f["pt_total"], f["pt_parcial"],
                   f["pt_iniciado"], f["pt_no_iniciado"], f["total_pct"]])
    ws.append([])
    ws.append(["Puntaje por nivel (% en Cumplimiento Total)"])
    ws[ws.max_row][0].font = negrita
    ws.append(["Nivel", "Gestión", "Puntaje %"])
    for c in ws[ws.max_row]:
        c.font = negrita
    for n in filas_nivel:
        ws.append([n["nivel"], n["nombre"], n["puntaje"]])
    ws.append([])
    ws.append(["% General", general])
    ws[ws.max_row][0].font = negrita
    ws.column_dimensions["C"].width = 45
    for col in ("A", "B", "D", "E", "F", "G", "H", "I"):
        ws.column_dimensions[col].width = 11

    ws2 = wb.create_sheet("Detalle 252")
    ws2.append(["Nivel", "Principio", "Requisito", "Código", "Elemento", "¿Cumple?",
                "Nivel de Cumplimiento", "Evidencia actual", "Evidencia anterior",
                "Comentario", "Acción de mejora", "Responsable", "Fecha"])
    for c in ws2[1]:
        c.font = negrita
    for v in listar_verificaciones(institucion_id):
        ws2.append([v["nivel_codigo"], f"{v['principio_numero']}. {v['principio_nombre']}",
                    v["requisito_codigo"], v["codigo"], v["texto"], v["cumple"] or "",
                    v["nivel_cumplimiento"] or "No Iniciado", v["evidencia_actual"] or "",
                    v["evidencia_anterior"] or "", v["comentario"] or "", v["accion_mejora"] or "",
                    v["responsable"] or "", v["fecha_cumplimiento"] or ""])
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["E"].width = 60

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _latin1(texto) -> str:
    return str(texto if texto is not None else "").encode("latin-1", "replace").decode("latin-1")


def exportar_resultados_pdf(institucion_id: int) -> bytes:
    from fpdf import FPDF

    filas, filas_nivel, general = _filas_resultados(institucion_id)
    nombre_inst = _nombre_institucion(institucion_id)
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 9, _latin1("Autodiagnóstico CNBPEO — Resultados"), ln=1)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, _latin1(f"Institución: {nombre_inst}"), ln=1)
    pdf.set_font("Helvetica", "", 10)
    fecha = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.cell(0, 6, _latin1(f"Código Nacional de Buenas Prácticas para las Estadísticas Oficiales · Generado: {fecha}"), ln=1)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, _latin1(f"% General de cumplimiento: {general if general is not None else '—'}%"), ln=1)
    pdf.ln(1)

    cols = [("Niv.", 12), ("N°", 8), ("Principio", 92), ("Tot", 12), ("Par", 12),
            ("Ini", 12), ("NoIni", 14), ("Falt", 12), ("Pautas", 16), ("% C.Total", 18)]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(0, 47, 108)
    pdf.set_text_color(255, 255, 255)
    for titulo, w in cols:
        pdf.cell(w, 7, _latin1(titulo), border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for f in filas:
        principio = f["principio"]
        if len(principio) > 62:
            principio = principio[:60] + "…"
        valores = [f["nivel"], str(f["numero"]), principio, str(f["total"]), str(f["parcial"]),
                   str(f["iniciado"]), str(f["no_iniciado"]), str(f["faltantes"]), str(f["total_pautas"]),
                   (f"{f['pt_total']}%" if f["pt_total"] is not None else "—")]
        for (titulo, w), val in zip(cols, valores):
            pdf.cell(w, 6, _latin1(val), border=1, align=("L" if titulo == "Principio" else "C"))
        pdf.ln()

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, _latin1("Puntaje por nivel (% en Cumplimiento Total)"), ln=1)
    pdf.set_font("Helvetica", "", 9)
    for n in filas_nivel:
        punt = f"{n['puntaje']}%" if n["puntaje"] is not None else "—"
        pdf.cell(0, 6, _latin1(f"  {n['nivel']} · {n['nombre']}: {punt}"), ln=1)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Material de apoyo (documentos subidos por el admin, guardados en la BD)
# ---------------------------------------------------------------------------

_MAX_MATERIAL_BYTES = 15 * 1024 * 1024  # 15 MB por archivo


def listar_material() -> list[dict]:
    """Lista los documentos de apoyo (sin el contenido binario)."""
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        return [dict(r) for r in cursor.execute(
            "SELECT id, titulo, nombre_archivo, mime, fecha_subida, categoria, "
            "LENGTH(contenido) AS tamano FROM autodx_material ORDER BY id"
        ).fetchall()]
    finally:
        conn.close()


def obtener_material(material_id: int) -> dict | None:
    """Devuelve un documento con su contenido binario, para descargar."""
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    try:
        fila = cursor.execute(
            "SELECT nombre_archivo, mime, contenido FROM autodx_material WHERE id = ?", (material_id,)
        ).fetchone()
        return dict(fila) if fila else None
    finally:
        conn.close()


def guardar_material(titulo: str, nombre_archivo: str, mime: str | None,
                     contenido: bytes, usuario_id: int | None,
                     categoria: str | None = None) -> tuple[bool, str]:
    """Guarda un documento de apoyo en la BD, en una categoría."""
    titulo = (titulo or "").strip() or (nombre_archivo or "Documento")
    if not contenido:
        return False, "El archivo está vacío."
    if len(contenido) > _MAX_MATERIAL_BYTES:
        return False, f"El archivo es muy grande (máx. {_MAX_MATERIAL_BYTES // (1024*1024)} MB)."
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        fecha = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
        cursor.execute(
            "INSERT INTO autodx_material (titulo, nombre_archivo, mime, contenido, subido_por, "
            "fecha_subida, categoria) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (titulo, nombre_archivo, mime, contenido, usuario_id, fecha, (categoria or "").strip() or None),
        )
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_MATERIAL_SUBIR", f"Material de apoyo: '{titulo}'.")
        conn.commit()
        return True, f"'{titulo}' subido correctamente."
    except sqlite3.Error as exc:
        conn.rollback()
        return False, f"No se pudo subir: {exc}"
    finally:
        conn.close()


def eliminar_material(material_id: int, usuario_id: int | None) -> tuple[bool, str]:
    _asegurar_esquema_v2()
    conn = db_mod.obtener_conexion()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM autodx_material WHERE id = ?", (material_id,))
        from models.logs import registrar_log
        registrar_log(cursor, usuario_id, "AUTODX_MATERIAL_ELIMINAR", f"Material de apoyo id={material_id}.")
        conn.commit()
        return True, "Documento eliminado."
    except sqlite3.Error as exc:
        conn.rollback()
        return False, f"No se pudo eliminar: {exc}"
    finally:
        conn.close()
